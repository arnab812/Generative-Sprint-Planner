# Generative Sprint Planner  

# TODO: imports

import streamlit as st
import pandas as pd
import math
from openpyxl.styles import Font
from datetime import datetime, timedelta

# alignment
from openpyxl.styles import Font, Alignment

# ! enabling download of the excel report in cloud
import io

# TODO: webpage configuration

# Set Streamlit page configuration for wider layout
st.set_page_config(layout="wide")

# TODO: team data

# Define team data
team_data = {
    "Agile Avengers": [
        {"Developer": "John Doe", "Role": "Associate Engineer-II", "Base Story Points": 6},
        {"Developer": "Finnian Waverly", "Role": "Engineer-I", "Base Story Points": 8},
        {"Developer": "Luna Thistledown", "Role": "Engineer-I", "Base Story Points": 30},
        {"Developer": "Jasper Voss", "Role": "Engineer-II", "Base Story Points": 10},
        {"Quality Analyst": "Cleo Marigold", "Role": "Quality Analyst-II", "Base Story Points": 30},
        {"Product Owner": "Orion Fable", "Role": "Product Manager-I"},
        {"Product Owner": "Tamsin Brightwood", "Role": "Product Manager-I"}
    ],
    "Scrumtastic Squad": [
        {"Developer": "Kieran Frost", "Role": "Senior Engineer-I", "Base Story Points": 12},
        {"Developer": "Elowen Starling", "Role": "Engineer-I", "Base Story Points": 8},
        {"Developer": "Bodhi Ember", "Role": "Senior Engineer-II", "Base Story Points": 14},
        {"Quality Analyst": "Zara Quillan", "Role": "Quality Analyst-I", "Base Story Points": 20},
        {"Quality Analyst": "Peter Haulet", "Role": "Quality Analyst-II", "Base Story Points": 30},
        {"Product Owner": "Johny English", "Role": "Product Manager-I"}
    ]
}

# TODO: impact table 

# Define story points impact table based on leaves
impact_table = {
    "Associate Engineer-II": [6.0, 5.4, 4.8, 4.2, 3.6, 3.0, 2.4, 1.8, 1.2, 0.6, 0.0],
    "Engineer-I": [8.0, 7.2, 6.4, 5.6, 4.8, 4.0, 3.2, 2.4, 1.6, 0.8, 0.0],
    "Engineer-II": [10.0, 9.0, 8.0, 7.0, 6.0, 5.0, 4.0, 3.0, 2.0, 1.0, 0.0],
    "Senior Engineer-I": [12.0, 10.8, 9.6, 8.4, 7.2, 6.0, 4.8, 3.6, 2.4, 1.2, 0.0],
    "Senior Engineer-II": [14.0, 12.6, 11.2, 9.8, 8.4, 7.0, 5.6, 4.2, 2.8, 1.4, 0.0],
    "Quality Analyst-I": [20.0, 18.0, 16.0, 14.0, 12.0, 10.0, 8.0, 6.0, 4.0, 2.0, 0.0],
    "Quality Analyst-II": [30.0, 27.0, 24.0, 21.0, 18.0, 15.0, 12.0, 9.0, 6.0, 3.0, 0.0]
}

# TODO: render.start() - components displayed on the UI

# * Header for the webpage
# Custom CSS for styling
st.markdown(
    """
    <style>
    .title {
        color: #85EBDF;
        font-family: monospace;
        font-size: 3em;
        text-align: left;
        transition: color 0.3s ease;
    }
    .title:hover {
        color: #65EFDF; /* Change color on hover */
        text-shadow: 2px 2px 5px rgba(0, 0, 0, 0.5); /* Add shadow effect */
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Title with custom class
st.markdown('<h1 class="title">Generative Sprint Planner</h1>', unsafe_allow_html=True)

# TODO: Sprint 
# Function to calculate sprint number based on the current date
def calculate_sprint_number(current_date):
    # Sprint starts on Friday and ends on next Thursday (2-week cycle)
    start_of_year = datetime(2024, 1, 5)  # First sprint starts on 5th January 2024 (Friday)
    
    # Calculate days difference between current date and the start of the year
    days_diff = (current_date - start_of_year).days
    
    # A sprint is 14 days long (2 weeks)
    sprint_number = (days_diff // 14) + 1  # Add 1 because we start counting from sprint 1
    return sprint_number

# Get the current date
current_date = datetime.today()

# Calculate the sprint number
sprint_number = calculate_sprint_number(current_date)

# Display the current sprint number
st.write(f"Sprint: {sprint_number + 1}")

# * Team selection
# team_selected = st.selectbox("Please select the scrum team", ["Agile Avengers", "Scrumtastic Squad"])
# Custom CSS for enhanced styling
st.markdown(
    """
    <style>
    .stSelectbox {
        background-color: #073642; /* Secondary background color */
        color: #E2EAF4; /* Text color */
        border: 2px solid #65B3EF; /* Border color */
        border-radius: 8px; /* Rounded corners */
        padding: 10px; /* Padding for better spacing */
        font-family: monospace; /* Font style */
        font-size: 1.2em; /* Increased font size */
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.2); /* Shadow effect */
        transition: border-color 0.3s ease, background-color 0.3s ease, transform 0.3s ease;
    }
    .stSelectbox:hover {
        border-color: #ED65EF; /* Change border color on hover */
        background-color: #002b36; /* Darker background on hover */
        transform: scale(1.02); /* Slightly enlarge on hover */
    }
    .stSelectbox:focus {
        outline: none; /* Remove default outline */
        border-color: #65B3EF; /* Highlight border on focus */
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Select box for team selection
team_selected = st.selectbox(
    "Please select the scrum team",
    ["Agile Avengers", "Scrumtastic Squad"]
)

# Get the selected team data
team_members = team_data[team_selected]

# ! Product Manager
# Function to get all Product Owners' names
def get_product_owners(team):
    product_owners = [member["Product Owner"] for member in team_data[team] if "Product Owner" in member]
    return product_owners
    
# Function to get initials for a name
def get_initials(name):
    initials = "".join([n[0] for n in name.split()]).upper()
    return initials

# Create empty DataFrame to collect final display data
developer_data = []
developer_data_for_snapshot = []
quality_analyst_data = []
quality_analyst_data_for_snapshot = []
product_owner_data = []
product_owner_data_for_snapshot = []

st.markdown("### Leave Counters")

col1, col2, col3 = st.columns(3)

# Process members based on their role

#with col2:
#    st.markdown("### Developers")
st.markdown(
    """
        <style>
            .dev {
                border: 2px solid #D5D036;
                padding-left: 8px;
                border-radius: 5px;
            }
            .prod {
                border: 2px solid #36D556;
                padding-left: 8px;
                border-radius: 5px;
            }
            .qa {
                border: 2px solid #DD2472;
                padding-left: 8px;
                border-radius: 5px;
            }
        </style>
    """, unsafe_allow_html=True
)
with col2:
    st.markdown('<h4 class="dev">Developers</h4>', unsafe_allow_html=True)

# with col3: 
#    st.markdown("### Quality Analysts")
with col3:
    st.markdown('<h4 class="qa">Quality Analysts</h4>', unsafe_allow_html=True)

# with col1:
#    st.markdown("### Product Owners")
with col1:
    st.markdown('<h4 class="prod">Product Owners</h4>', unsafe_allow_html=True)

for member in team_members:
    if "Developer" in member:
        with col2:
            initials = get_initials(member["Developer"])
            name_with_icon = f"""
            <div style="display: flex; align-items: center;">
                <div style="background-color: #0073e6; color: white; border-radius: 50%; width: 30px; height: 30px; 
                            display: flex; align-items: center; justify-content: center; font-weight: bold; margin-right: 8px;">
                    {initials}
                </div>
                <span>{member["Developer"]}</span>
            </div>
            """
            dev_leave = st.number_input(f"Leave(s) for {member['Developer']}", min_value=0, max_value=10, value=0, step=1, key=member["Developer"])
            available_hours = (10 - dev_leave) * 8
            points = impact_table[member["Role"]][dev_leave]
            
            developer_data.append([name_with_icon, member["Role"], dev_leave, available_hours, points])

            # ***** new 
            developer_data_for_snapshot.append([member['Developer'], member["Role"], dev_leave, available_hours, points])

    elif "Quality Analyst" in member:
        with col3:
            qa_name = member["Quality Analyst"]
            qa_role = member["Role"]
            
            # Get initials for QA
            qa_initials = get_initials(qa_name)
            qa_name_with_icon = f"""
            <div style="display: flex; align-items: center;">
                <div style="background-color: #0073e6; color: white; border-radius: 50%; width: 30px; height: 30px; 
                            display: flex; align-items: center; justify-content: center; font-weight: bold; margin-right: 8px;">
                    {qa_initials}
                </div>
                <span>{qa_name}</span>
            </div>
            """
            
            qa_leave = st.number_input(f"Leave(s) for {qa_name}", min_value=0, max_value=10, value=0, step=1, key=qa_name)
            qa_available_hours = (10 - qa_leave) * 8
            qa_points = impact_table[qa_role][qa_leave]
            
            quality_analyst_data.append([qa_name_with_icon, qa_role, qa_leave, qa_available_hours, qa_points])

            # **** new
            quality_analyst_data_for_snapshot.append([member['Quality Analyst'], qa_role, qa_leave, qa_available_hours, qa_points])

    elif "Product Owner" in member:
        with col1:
            initials = get_initials(member["Product Owner"])
            name_with_icon = f"""
            <div style="display: flex; align-items: center;">
                <div style="background-color: #0073e6; color: white; border-radius: 50%; width: 30px; height: 30px; 
                            display: flex; align-items: center; justify-content: center; font-weight: bold; margin-right: 8px;">
                    {initials}
                </div>
                <span>{member["Product Owner"]}</span>
            </div>
            """
            po_leave = st.number_input(f"Leave(s) for {member['Product Owner']}", min_value=0, max_value=10, value=0, step=1, key=member["Product Owner"])
            po_available_hours = (10 - po_leave) * 8
            product_owner_data.append([name_with_icon, member["Role"], po_leave, po_available_hours])
            product_owner_data_for_snapshot.append([member['Product Owner'], member["Role"], po_leave, po_available_hours])

# TODO: 1. Display Product Owners Table
st.markdown("### Product Team")
st.markdown("""
<div id="prod-table" style="display: grid; grid-template-columns: 2fr 1fr 1fr 1fr 2fr; font-weight: bold; background-color: #073642; border: 2px solid #36D556; border-radius: 5px; padding: 10px; margin-bottom: 10px; color: #E2EAF4; font-family: monospace;">
    <div style="flex: 2;">Product Owner</div>
    <div style="flex: 1; text-align: center;">Role</div>
    <div style="flex: 1; text-align: center;">Leave(day(s))</div>
    <div style="flex: 1; text-align: center;">Available Hours(h)</div>
</div>
""", unsafe_allow_html=True)
for row in product_owner_data:
    st.markdown(f"""
    <div id="prod-table" style="display: grid; grid-template-columns: 2fr 1fr 1fr 1fr 2fr; font-weight: bold; background-color: #073642; border: 2px solid #36D556; border-radius: 5px; padding: 10px; margin-bottom: 10px; color: #E2EAF4; font-family: monospace;">
        <div style="flex: 2;">{row[0]}</div>
        <div style="flex: 1; text-align: center;">{row[1]}</div>
        <div style="flex: 1; text-align: center;">{row[2]}</div>
        <div style="flex: 1; text-align: center;">{row[3]}</div>
    </div>
    """, unsafe_allow_html=True)

# Add hover effect to each row using CSS
st.markdown("""
<style>
#prod-table div:hover {
    background-color: #004d62;
    cursor: pointer;
    transform: scale(1.02);
    transition: transform 0.3s ease;
    font-weight: bold;
}
</style>
""", unsafe_allow_html=True)

# TODO: 2. Display Developers Table
st.markdown("### Development Team")

# Header with updated design and grid layout
st.markdown(f"""
<div id="dev-table" style="display: grid; grid-template-columns: 2fr 1fr 1fr 1fr 2fr; font-weight: bold; background-color: #073642; border: 2px solid #D5D036; border-radius: 5px; padding: 10px; margin-bottom: 10px; color: #E2EAF4; font-family: monospace;">
    <div style="text-align: left; font-size: 16px;">Developer</div>
    <div style="text-align: center; font-size: 16px;">Role</div>
    <div style="text-align: center; font-size: 16px;">Leave(day(s))</div>
    <div style="text-align: center; font-size: 16px;">Available Hours(h)</div>
    <div style="text-align: center; font-size: 16px;">Accommodatable Story Points</div>
</div>
""", unsafe_allow_html=True)

# Rows with grid layout, hover effect, and theme-based coloring
for idx, row in enumerate(developer_data):
    row_id = f"row-{idx}"  # Unique ID for each row
    st.markdown(f"""
    <div id="dev-table" style="display: grid; grid-template-columns: 2fr 1fr 1fr 1fr 2fr; background-color: #002b36; border: 1px solid #D5D036; border-radius: 5px; padding: 10px; margin-bottom: 5px; transition: background-color 0.3s; color: #E2EAF4; font-family: monospace;">
        <div style="text-align: left; font-size: 14px;">{row[0]}</div>
        <div style="text-align: center; font-size: 14px;">{row[1]}</div>
        <div style="text-align: center; font-size: 14px;">{row[2]}</div>
        <div style="text-align: center; font-size: 14px;">{row[3]}</div>
        <div style="text-align: center; font-size: 14px;">{int(row[4])}</div>
    </div>
    """, unsafe_allow_html=True)

# Add hover effect to each row using CSS
st.markdown("""
<style>
#dev-table div:hover {
    background-color: #004d62;
    cursor: pointer;
    transform: scale(1.02);
    transition: transform 0.3s ease;
    font-weight: bold;
}
</style>
""", unsafe_allow_html=True)

# TODO 3. Display Quality Analysts Table
st.markdown("### Quality Assurance Team")
st.markdown(""" 
<div id="qa-table" style="display: grid; grid-template-columns: 2fr 1fr 1fr 1fr 2fr; font-weight: bold; background-color: #073642; border: 2px solid #DD2472; border-radius: 5px; padding: 10px; margin-bottom: 10px; color: #E2EAF4; font-family: monospace;"> 
    <div style="flex: 2;">Quality Analyst</div>
    <div style="flex: 1; text-align: center">Role</div>
    <div style="flex: 1; text-align: center;">Leave(day(s))</div>
    <div style="flex: 1; text-align: center;">Available Hours(h)</div>
    <div style="flex: 2; text-align: center;">Accommodatable Story Points</div>
</div>
""", unsafe_allow_html=True)

for row in quality_analyst_data:
    st.markdown(f"""
    <div id="qa-table" style="display: grid; grid-template-columns: 2fr 1fr 1fr 1fr 2fr; font-weight: bold; background-color: #073642; border: 2px solid #DD2472; border-radius: 5px; padding: 10px; margin-bottom: 10px; color: #E2EAF4; font-family: monospace;">
        <div style="flex: 2;">{row[0]}</div>
        <div style="flex: 1; text-align: center;">{row[1]}</div>
        <div style="flex: 1; text-align: center;">{row[2]}</div>
        <div style="flex: 1; text-align: center;">{row[3]}</div>
        <div style="flex: 2; text-align: center;">{int(row[4])}</div>
    </div>
    """, unsafe_allow_html=True)

# Add hover effect to each row using CSS
st.markdown("""
<style>
#qa-table div:hover {
    background-color: #004d62;
    cursor: pointer;
    transform: scale(1.02);
    transition: transform 0.3s ease;
    font-weight: bold;
}
</style>
""", unsafe_allow_html=True)

# Calculate and display total accommodatable story points excluding Quality Analysts
total_story_points = math.floor(sum(row[4] for row in developer_data))  # Only sum developer story points
st.write(f"**Total Accommodatable Story Points: {total_story_points}**")

# Add condition to calculate required DevTesting
qa_story_points = math.floor(sum(row[4] for row in quality_analyst_data))

if total_story_points > qa_story_points:
    required_dev_testing = total_story_points - qa_story_points
    if required_dev_testing > 1:
        st.write(f"**DevTesting may be Required of Story Points: {required_dev_testing}**")
    elif required_dev_testing == 1:
        st.write(f"**DevTesting may be Required of Story Point: 1**")

#! -- #

# Convert Data to DataFrames for Snapshot
developer_df = pd.DataFrame(developer_data_for_snapshot, columns=["Developer", "Role", "Leave(day(s))", "Available Hours(h)", "Story Points"])
quality_analyst_df = pd.DataFrame(quality_analyst_data_for_snapshot, columns=["Quality Analyst", "Role", "Leave(day(s))", "Available Hours(h)", "Story Points"])
product_owner_df = pd.DataFrame(product_owner_data_for_snapshot, columns=["Product Owner", "Role", "Leave(day(s))", "Available Hours(h)"])

# Capture Snapshot Button
if st.button("Generate Sprint Snapshot"):
    # Use BytesIO to create an in-memory buffer
    with io.BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # Convert data to DataFrames
            developer_df = pd.DataFrame(developer_data_for_snapshot, columns=["Developer", "Role", "Leave(day(s))", "Available Hours(h)", "Story Points"])
            quality_analyst_df = pd.DataFrame(quality_analyst_data_for_snapshot, columns=["Quality Analyst", "Role", "Leave(day(s))", "Available Hours(h)", "Story Points"])
            product_owner_df = pd.DataFrame(product_owner_data_for_snapshot, columns=["Product Owner", "Role", "Leave(day(s))", "Available Hours(h)"])

            # Write Developers data to the sheet
            developer_df.to_excel(writer, sheet_name="Snapshot", index=False, startrow=0)
            
            # Write Quality Analysts data below Developers data
            quality_startrow = len(developer_df) + 2  # Add 2 rows for spacing
            quality_analyst_df.to_excel(writer, sheet_name="Snapshot", index=False, startrow=quality_startrow)
            
            # Write Product Owners data below Quality Analysts data
            product_owner_startrow = quality_startrow + len(quality_analyst_df) + 2  # Add 2 rows for spacing
            product_owner_df.to_excel(writer, sheet_name="Snapshot", index=False, startrow=product_owner_startrow)
            
            # Metrics start row (below Product Owners data)
            metrics_startrow = product_owner_startrow + len(product_owner_df) + 3
            
            # Add column headers "Metric" and "Value"
            worksheet = writer.sheets["Snapshot"]
            worksheet.cell(row=metrics_startrow, column=1).value = "Metric"
            bold_font = Font(bold=True)
            worksheet.cell(row=metrics_startrow, column=1).font = bold_font
            worksheet.cell(row=metrics_startrow, column=1).alignment = center_alignment
            
            # Add "Total Accommodatable Story Points" and its value
            total_story_points = sum([data[-1] for data in developer_data_for_snapshot])  # Example calculation
            worksheet.cell(row=metrics_startrow + 1, column=1).value = f"Total Accommodatable Story Points: {total_story_points}"
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 45
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 20
        
        # Get the value of the in-memory buffer
        buffer.seek(0)
        excel_data = buffer.getvalue()

    # Create a download button
    st.download_button(
        label="Download Sprint Snapshot",
        data=excel_data,
        file_name=f"{team_selected} Sprint-{sprint_number+1} Snapshot.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.success(f"{team_selected} Sprint-{sprint_number+1} Snapshot prepared successfully!")


# TODO: About me
st.markdown("""
<div style="padding-top: 50px;">
    <h3>About me building Generative Sprint Planner</h3>
    <p>
        As a passionate <strong>Software Engineer</strong>, I thrive on crafting solutions that streamline workflows and address real-world challenges. My journey specific to this <strong>Sprint Planner</strong> has been fueled by curiosity and a commitment to creating a tool that makes <u style="color: yellow">sprint planning</u> and <u style="color: yellow">impact calculation</u> <em style="color: #70FA64">automated without requiring any additional meeting</em>. With a solid foundation in <strong>Agile methodologies</strong>, <strong>Scrum practices</strong>, and <strong>Sprint management</strong>, I set out to tackle this repetitive task in an <u style="color: yellow">efficient</u> and <u style="color: yellow">auto-generated manner</u>.
    </p>
    <p>
        The purpose of this application is to help the engineering community save valuable time by <u style="color: yellow">automating impact calculations during sprint planning</u>. 
        While it may serve a smaller purpose compared to comprehensive tools like <strong>Jira</strong>(I personally love the most, as my daily-driver), I am excited about 
        the opportunity to make a meaningful impact on other developers' lives through my work.
    </p>
    <p>
        This app is a testament to my belief that even the smallest contributions can spark significant improvements in the way we work as a community. <br><br> Thank you!
    </p>
</div>
""", unsafe_allow_html=True)

# Path to the image file
image_path = "Arnab_Chakraborty_Buidling_Generative_Sprint_Planner.jpg"

st.image(image_path, caption="", width=200)

# Display clickable LinkedIn link
st.markdown(
    """
    <div style="padding-left: 20px">
        <a href="https://www.linkedin.com/in/arnab-chakraborty-2881a31b7/" target="_blank">
            Arnab Chakraborty
        </a>
    </div>
    """,
    unsafe_allow_html=True
)
